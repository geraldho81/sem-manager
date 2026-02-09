from pathlib import Path
from typing import Dict, Any, List
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.models.rsa import MediaPlan
from app.config import MARKETS


# -- Style constants --
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SUBHEADER_FILL = PatternFill("solid", fgColor="D6E4F0")
SUBHEADER_FONT = Font(bold=True, color="1F4E79", size=10)
LABEL_FONT = Font(bold=True, color="1F4E79", size=10)
DATA_FONT = Font(color="333333", size=10)
MONEY_FONT = Font(color="0000FF", size=10)  # Blue = hardcoded inputs
TOTAL_FILL = PatternFill("solid", fgColor="E2EFDA")
TOTAL_FONT = Font(bold=True, color="1F4E79", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)


def _style_header_row(ws, row, col_start, col_end):
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def _style_data_row(ws, row, col_start, col_end, fill=None):
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = DATA_FONT
        cell.alignment = LEFT
        cell.border = THIN_BORDER
        if fill:
            cell.fill = fill


def _auto_width(ws, min_width=10, max_width=40):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        width = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = width


class ExcelExporter:
    """Exports media plans to Excel workbooks matching agency format."""

    async def export(
        self,
        media_plan: MediaPlan,
        strategy: Dict[str, Any],
        keyword_research: Dict[str, Any],
        brand_research: Dict[str, Any],
        output_folder: str,
    ) -> str:
        output_path = Path(output_folder)
        output_path.mkdir(parents=True, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"media_plan_{timestamp}.xlsx"
        filepath = output_path / filename

        wb = Workbook()

        market_config = MARKETS.get(media_plan.market, {})
        currency = market_config.get("currency", "USD")
        currency_symbol = market_config.get("currency_symbol", "$")
        market_name = market_config.get("name", media_plan.market)
        brand_name = brand_research.get("brand_name", "Brand")
        campaign_name = media_plan.campaign_name

        # Sheet 1: Media Plan
        ws_plan = wb.active
        ws_plan.title = "SEM - Media Plan"
        self._build_media_plan_sheet(
            ws_plan, media_plan, strategy, brand_name, campaign_name,
            market_name, currency, currency_symbol,
        )

        # Sheet 2: Keywords
        ws_kw = wb.create_sheet("SEM - Keywords")
        self._build_keywords_sheet(
            ws_kw, media_plan, keyword_research, campaign_name,
            currency, currency_symbol,
        )

        # Sheet 3: RSA Ads
        ws_rsa = wb.create_sheet("SEM - RSA Ads")
        self._build_rsa_sheet(ws_rsa, media_plan, campaign_name)

        # Sheet 4: Strategy Overview
        ws_strat = wb.create_sheet("Strategy Overview")
        self._build_strategy_sheet(ws_strat, strategy, brand_name, market_name)

        wb.save(str(filepath))
        return str(filepath)

    # ------------------------------------------------------------------
    # Sheet 1: Media Plan
    # ------------------------------------------------------------------
    def _build_media_plan_sheet(
        self, ws, media_plan, strategy, brand_name, campaign_name,
        market_name, currency, currency_symbol,
    ):
        # Header info block
        info = [
            ("Client", brand_name),
            ("Campaign", campaign_name),
            ("Platform", "Google Ads"),
            ("Geo Location", market_name),
            ("Currency", currency),
            ("Landing Page", media_plan.landing_page_urls[0] if media_plan.landing_page_urls else ""),
            ("Date", datetime.now().strftime("%d %b %Y")),
        ]

        row = 1
        for label, value in info:
            ws.cell(row=row, column=1, value=label).font = LABEL_FONT
            ws.cell(row=row, column=2, value=value).font = DATA_FONT
            row += 1

        row += 1  # blank separator

        # Table headers
        headers = [
            "Platform", "Campaign Type", "Campaign Name", "Ad Group",
            "Targeting / Theme", "Creative Format", "KPI",
            "Geo Location", "Keywords", f"Avg CPC ({currency_symbol})",
            f"Monthly Budget ({currency_symbol})", f"Daily Budget ({currency_symbol})",
        ]
        for col, h in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=h)
        _style_header_row(ws, row, 1, len(headers))
        header_row = row
        row += 1

        # One row per ad group
        num_cols = len(headers)
        for ag in media_plan.ad_groups:
            kw_count = len(ag.keywords)
            kw_list = ", ".join(k.text for k in ag.keywords[:5])
            if kw_count > 5:
                kw_list += f" (+{kw_count - 5} more)"

            avg_cpc = ag.cpc_bid
            if ag.keywords:
                cpcs = [k.cpc for k in ag.keywords if k.cpc]
                if cpcs:
                    avg_cpc = sum(cpcs) / len(cpcs)

            total_volume = sum(k.monthly_volume or 0 for k in ag.keywords)
            est_monthly = round(avg_cpc * total_volume * 0.03, 2) if total_volume else round(avg_cpc * kw_count * 30, 2)
            est_daily = round(est_monthly / 30, 2)

            values = [
                "Google Ads",
                "Search",
                campaign_name,
                ag.name,
                ag.name,  # theme = ad group name
                "RSA",
                "CPA / ROAS",
                market_name,
                kw_list,
                round(avg_cpc, 2),
                est_monthly,
                est_daily,
            ]
            for col, v in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=v)
                cell.font = DATA_FONT
                cell.border = THIN_BORDER
                cell.alignment = LEFT
                # Money columns in blue
                if col in (10, 11, 12):
                    cell.font = MONEY_FONT

            row += 1

        # Totals row
        data_start = header_row + 1
        data_end = row - 1
        ws.cell(row=row, column=1, value="TOTAL").font = TOTAL_FONT
        for col in range(1, num_cols + 1):
            ws.cell(row=row, column=col).fill = TOTAL_FILL
            ws.cell(row=row, column=col).border = THIN_BORDER
            ws.cell(row=row, column=col).font = TOTAL_FONT
        # Sum formulas for budget columns
        col_monthly = 11
        col_daily = 12
        ws.cell(row=row, column=col_monthly, value=f"=SUM({get_column_letter(col_monthly)}{data_start}:{get_column_letter(col_monthly)}{data_end})")
        ws.cell(row=row, column=col_daily, value=f"=SUM({get_column_letter(col_daily)}{data_start}:{get_column_letter(col_daily)}{data_end})")

        _auto_width(ws, min_width=12, max_width=45)

    # ------------------------------------------------------------------
    # Sheet 2: Keywords
    # ------------------------------------------------------------------
    def _build_keywords_sheet(
        self, ws, media_plan, keyword_research, campaign_name,
        currency, currency_symbol,
    ):
        # Build lookup of keyword data from keyword_research clusters
        kw_lookup: Dict[str, Dict] = {}
        for cluster in keyword_research.get("clusters", []):
            for kw in cluster.get("keywords", []):
                keyword_text = kw.get("keyword", "").lower().strip()
                kw_lookup[keyword_text] = kw

        # Headers
        headers = [
            "Campaign", "Ad Group", "Criterion Type", "Keyword",
            f"Search Volume", f"Competition",
            f"Low Bid ({currency_symbol})", f"High Bid ({currency_symbol})",
            "Character Count", "Match Type", "Status",
        ]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
        _style_header_row(ws, 1, 1, len(headers))

        row = 2
        num_cols = len(headers)

        for ag in media_plan.ad_groups:
            # Ad group sub-header row
            ws.cell(row=row, column=1, value=campaign_name)
            ws.cell(row=row, column=2, value=ag.name)
            for col in range(1, num_cols + 1):
                cell = ws.cell(row=row, column=col)
                cell.fill = SUBHEADER_FILL
                cell.font = SUBHEADER_FONT
                cell.border = THIN_BORDER
            row += 1

            for kw in ag.keywords:
                kw_text = kw.text
                kw_data = kw_lookup.get(kw_text.lower().strip(), {})

                search_vol = kw.monthly_volume or kw_data.get("search_volume", "")
                competition = kw_data.get("competition", "")
                if isinstance(competition, float) and competition > 0:
                    if competition < 0.33:
                        competition_label = "Low"
                    elif competition < 0.67:
                        competition_label = "Medium"
                    else:
                        competition_label = "High"
                elif isinstance(competition, str) and competition:
                    competition_label = competition
                else:
                    competition_label = ""

                cpc = kw.cpc or kw_data.get("cpc", 0)
                low_bid = round(cpc * 0.7, 2) if cpc else ""
                high_bid = round(cpc * 1.3, 2) if cpc else ""
                match_type = kw.match_type.capitalize() if kw.match_type else "Broad"

                values = [
                    campaign_name,
                    ag.name,
                    "Keyword",
                    kw_text,
                    search_vol if search_vol else "",
                    competition_label,
                    low_bid,
                    high_bid,
                    len(kw_text),
                    match_type,
                    "Enabled",
                ]
                for col, v in enumerate(values, 1):
                    cell = ws.cell(row=row, column=col, value=v)
                    cell.font = DATA_FONT
                    cell.border = THIN_BORDER
                    cell.alignment = LEFT
                    if col in (7, 8):
                        cell.font = MONEY_FONT

                row += 1

            # Blank separator between ad groups
            row += 1

        _auto_width(ws, min_width=10, max_width=50)

    # ------------------------------------------------------------------
    # Sheet 3: RSA Ads
    # ------------------------------------------------------------------
    def _build_rsa_sheet(self, ws, media_plan, campaign_name):
        headers = [
            "Campaign", "Ad Group", "Type",
            "Headline 1", "Headline 2", "Headline 3", "Headline 4", "Headline 5",
            "Headline 6", "Headline 7", "Headline 8", "Headline 9", "Headline 10",
            "Headline 11", "Headline 12", "Headline 13", "Headline 14", "Headline 15",
            "Description 1", "Description 2", "Description 3", "Description 4",
            "Final URL", "Status",
        ]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
        _style_header_row(ws, 1, 1, len(headers))

        row = 2
        landing_url = media_plan.landing_page_urls[0] if media_plan.landing_page_urls else ""
        num_cols = len(headers)

        for ag in media_plan.ad_groups:
            ws.cell(row=row, column=1, value=campaign_name)
            ws.cell(row=row, column=2, value=ag.name)
            ws.cell(row=row, column=3, value="Responsive Search Ad")

            for i in range(15):
                col_idx = 4 + i
                if i < len(ag.headlines):
                    ws.cell(row=row, column=col_idx, value=ag.headlines[i])

            for i in range(4):
                col_idx = 19 + i
                if i < len(ag.descriptions):
                    ws.cell(row=row, column=col_idx, value=ag.descriptions[i])

            ws.cell(row=row, column=23, value=landing_url)
            ws.cell(row=row, column=24, value="Enabled")

            for col in range(1, num_cols + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = DATA_FONT
                cell.border = THIN_BORDER
                cell.alignment = WRAP if col >= 4 else LEFT

            row += 1

        _auto_width(ws, min_width=12, max_width=35)
        # Widen headline/description columns
        for col in range(4, 23):
            ws.column_dimensions[get_column_letter(col)].width = 30

    # ------------------------------------------------------------------
    # Sheet 4: Strategy Overview
    # ------------------------------------------------------------------
    def _build_strategy_sheet(self, ws, strategy, brand_name, market_name):
        row = 1
        info = [
            ("Brand", brand_name),
            ("Campaign", strategy.get("campaign_name", "")),
            ("Objective", strategy.get("objective", "")),
            ("Budget Recommendation", strategy.get("budget_recommendation", "")),
            ("Bidding Strategy", strategy.get("bidding_strategy", "")),
            ("Market", market_name),
            ("Targeting Notes", strategy.get("targeting_notes", "")),
        ]
        for label, value in info:
            ws.cell(row=row, column=1, value=label).font = LABEL_FONT
            ws.cell(row=row, column=2, value=str(value)).font = DATA_FONT
            row += 1

        row += 1

        # Negative keywords
        negatives = strategy.get("negative_keywords", [])
        if negatives:
            ws.cell(row=row, column=1, value="Negative Keywords").font = LABEL_FONT
            row += 1
            for nk in negatives:
                ws.cell(row=row, column=1, value=nk).font = DATA_FONT
                row += 1
            row += 1

        # Ad Group strategy details
        ws.cell(row=row, column=1, value="Ad Group Strategies").font = Font(bold=True, color="1F4E79", size=12)
        row += 1

        ag_headers = ["Ad Group", "Theme", "Target Persona", "Messaging Angle", "Priority", "Suggested Bid"]
        for col, h in enumerate(ag_headers, 1):
            ws.cell(row=row, column=col, value=h)
        _style_header_row(ws, row, 1, len(ag_headers))
        row += 1

        for ag in strategy.get("ad_groups", []):
            values = [
                ag.get("name", ""),
                ag.get("theme", ""),
                ag.get("target_persona", ""),
                ag.get("messaging_angle", ""),
                ag.get("priority", ""),
                ag.get("suggested_bid", ""),
            ]
            for col, v in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=v)
                cell.font = DATA_FONT
                cell.border = THIN_BORDER
                cell.alignment = LEFT
            row += 1

        _auto_width(ws, min_width=12, max_width=50)
